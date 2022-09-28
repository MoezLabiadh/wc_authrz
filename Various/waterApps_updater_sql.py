import os
import cx_Oracle
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


def connect_to_DB (username,password,hostname):
    """ Returns a connection to Oracle database"""
    try:
        connection = cx_Oracle.connect(username, password, hostname, encoding="UTF-8")
        print  ("Successffuly connected to the database")
    except:
        raise Exception('Connection failed! Please verifiy your login parameters')

    return connection

    
def read_query(connection,query):
    "Returns a df containing SQL Query results"
    cursor = connection.cursor()
    try:
        cursor.execute(query)
        names = [x[0] for x in cursor.description]
        rows = cursor.fetchall()
        return pd.DataFrame(rows, columns=names)
    
    finally:
        if cursor is not None:
            cursor.close()


def update_wtshd_info(f, connection, sql, df_lk, saveOp):
    """Updates the Watershed and Region info of the Water Apps ledger"""
    wb = load_workbook(f)
    ws = wb['Active Applications']
    
    for i, row in enumerate(ws.iter_rows(min_row=2,values_only=True)):
        row_id = i+2
    
        if ws['X{}'.format(row_id)].value is None:
            print ("\n********Working on Row {}********".format (row_id))
            lat = row[10]
            long = row[11]
            
            print ('Latitude: {}'. format(lat))
            print ('Longitude: {}'. format(long))
            
            if (lat is None) or (long is None):
                raise Exception("Latitude or/and Longitude values are not specified!")
                
            elif not min(54.465, 48.2) < lat < max(54.5, 48.134):
                raise Exception("Latitude value is out of range!")
                
            
            elif not min(-122.634, -133.257) < long < max(-122.7, -133.3):
                    raise Exception("Longitude value is out of range!")  
                
            else:
                query = sql.format (long=long, lat=lat)
                df_sql = read_query(connection,query)
                
                wtrsh_name = df_sql['WATER_LICENSING_WATERSHED_NAME'].iloc[0]
                ws['X{}'.format(row_id)] = wtrsh_name
                ws['Y{}'.format(row_id)]  = """=VLOOKUP(X{},'Pick Lists'!$L$1:$M$107,2,FALSE)""".format(row_id)
                
                df_lk_i = df_lk.loc[df_lk['WATER_LICENSING_WATERSHED'] == wtrsh_name]
                print ('Watershed: {}'.format(wtrsh_name))
                print ('Sub-region: {}'.format(df_lk_i['REGION'].iloc[0]))
            
        else:
            #raise Exception ('No Empty watershed values found in spreadsheet')
            pass
        
    #Add data validation for Watershed name.    
    dv = DataValidation(type="list",
                        formula1= "'Pick Lists'!$L$2:$L$107",
                        allow_blank=False,
                        showDropDown= False)
    
    dv.add("X2:X" + str(ws.max_row))
    ws.add_data_validation(dv)
    
    if saveOp == 'Yes':
        print('Writing Watershed and Subregion info to the Spreadsheet')
        wb.save(f)
        
    else:
        pass
    
def main():
    workspace = r'\\spatialfiles.bcgov\Work\lwbc\visr\Workarea\moez_labiadh\WORKSPACE\20220916_waterLicencing_support'
    f = os.path.join (workspace,'Water Application Ledger_workingCopy.xlsx')
    
    saveOp = 'No' # yes or no
    
    df_lk = pd.read_excel(f, 'Pick Lists')
    df_lk = df_lk[['WATER_LICENSING_WATERSHED', 'REGION']]
    
    print ('Connecting to BCGW...')
    hostname = 'bcgw.bcgov/idwprod1.bcgov'
    bcgw_user = os.getenv('bcgw_user')
    bcgw_pwd = os.getenv('bcgw_pwd')
    connection = connect_to_DB (bcgw_user,bcgw_pwd,hostname)
    
    sql = '''
            SELECT 
                WATER_LICENSING_WATERSHED_NAME
            FROM 
                WHSE_WATER_MANAGEMENT.WLS_WATER_LIC_WATERSHEDS_SP wsh
            WHERE 
                SDO_RELATE (wsh.SHAPE, 
                            SDO_GEOMETRY('POINT({long} {lat})', 4326),
                            'mask=ANYINTERACT') = 'TRUE'
        '''
        
    print ("Retrieving Watershed and Sub-region info...")
    update_wtshd_info(f, connection, sql, df_lk,saveOp)
    
    print ()
    
main()
