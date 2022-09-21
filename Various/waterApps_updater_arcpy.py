import os
import arcpy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


def update_wtshd_info(f, wtsh_fc,df_lk):
    """Updates the Watershed and Region info of the Water Apps ledger"""
    wb = load_workbook(f)
    ws = wb['Active Applications']

    for i, row in enumerate(ws.iter_rows(min_row=2,values_only=True)):
        row_id = i+2
        if ws['X{}'.format(row_id)].value is None:
            print ("..working on Row {}".format (row_id))
            lat = row[10]
            long = row[11]

            if (lat is None) or (long is None):
                raise Exception("ERROR: Latitude or/and Longitude values are not specified!")

            else:
                pt = arcpy.Point(float(long),float(lat))
                sr = arcpy.SpatialReference("WGS 1984")
                ptGeo = arcpy.PointGeometry(pt,sr)

                pt_fc = 'in_memory\pt_fc'
                arcpy.management.CopyFeatures(ptGeo, pt_fc)

                intr = 'in_memory\_intersect'
                arcpy.Intersect_analysis([pt_fc,wtsh_fc], intr)

                with arcpy.da.UpdateCursor(intr,['WATER_LICENSING_WATERSHED_NAME']) as cursor:
                    for row in cursor:
                        wtrsh_name = row[0]

                ws['X{}'.format(row_id)] = wtrsh_name
                ws['Y{}'.format(row_id)]  = """=VLOOKUP(X{},'Pick Lists'!$L$1:$M$107,2,FALSE)""".format(row_id)

                df_lk_i = df_lk.loc[df_lk['WATER_LICENSING_WATERSHED'] == wtrsh_name]
                print ('....Watershade is {}'.format(wtrsh_name))
                print ('....Region is {}'.format(df_lk_i['REGION'].iloc[0]))

                wb.save(f)

                arcpy.Delete_management('in_memory')

        else:
            pass

    #Add data validation for Watershed name.
    dv = DataValidation(type="list",
                        formula1= "'Pick Lists'!$L$2:$L$107",
                        allow_blank=False,
                        showDropDown= False)

    dv.add("X2:X" + str(ws.max_row))
    ws.add_data_validation(dv)

    wb.save(f)

def main():
    workspace = r'\\spatialfiles.bcgov\Work\lwbc\visr\Workarea\moez_labiadh\WORKSPACE\20220916_waterLicencing_support'
    f = os.path.join (workspace,'Water Application Ledger_workingCopy2.xlsx')
    wtsh_fc = os.path.join (workspace,'data.gdb', 'watersheds_WC_all')

    df_lk = pd.read_excel(f, 'Pick Lists')
    df_lk = df_lk[['WATER_LICENSING_WATERSHED', 'REGION']]

    print ("Updating Watershed and Region info...")
    update_wtshd_info(f, wtsh_fc,df_lk)

main()
