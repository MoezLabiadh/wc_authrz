
#-------------------------------------------------------------------------------
# Name:        Land Files Tracker
#
# Purpose:     This script generates the Monthly Land files 
#              tracking reports for West Coast Region: backlog and active files.
#
# Input(s):    (1) Titan workledger report RPT009 (excel) 
#              (2) ATS processing time report (detailed export).
#              (3) ATS on-hold authorzations report (spreadsheet export).
#              (4) ATS bring forward report (data-dump export).
# 
# Note: Place the obove input reports in the FILE_TRACKING\INPUTS folder
#   
# Author:      Moez Labiadh - FCBC, Nanaimo
#
# Created:     2024-05-01
# Updated:     2024-09-04
#-------------------------------------------------------------------------------

import warnings
warnings.simplefilter(action='ignore')

import os

import cx_Oracle
import pandas as pd
#import numpy as np

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

from datetime import date, timedelta

import plotly.express as px
import plotly.graph_objects as go

from PIL import Image as PILImage


def connect_to_DB (username,password,hostname):
    """ Returns a connection and cursor to Oracle database"""
    try:
        connection = cx_Oracle.connect(username, password, hostname, encoding="UTF-8")
        print  ("....Successffuly connected to the database")
    except:
        raise Exception('....Connection failed! Please check your login parameters')

    return connection


def import_ats_oh (ats_oh_f):
    """Reads the ATS Auth. On Hold report into a df"""
    df = pd.read_html(ats_oh_f)[5]
    df.columns = df.iloc[1]
    df.drop([0, 1],inplace=True)
    
    df['On Hold End Date']=''
    cols_onh = ['Project Number','On Hold Start Date', 
                'On Hold End Date','Reason For Hold']
    
    df = df[cols_onh]
    
    return df


def import_ats_bf (ats_bf_f):
    """Reads the ATS Bring Forward report into a df"""
    warnings.filterwarnings("ignore", category=pd.errors.ParserWarning)
    

    #df = pd.read_csv(ats_bf_f, delimiter="\t",encoding='cp1252',error_bad_lines=False)
    df = pd.read_csv(ats_bf_f, delimiter="\t",encoding='cp1252',on_bad_lines='skip')

    cols_onh = ['Project Number','Authorization Assigned To', 
                'Bring Forward Date']
    
    df = df[cols_onh]
    df.dropna(subset=['Project Number'],inplace=True)
    
    return df


def import_ats_pt (ats_pt_f, df_onh,df_bfw):
    """Reads the ATS Processing Time report into a df"""
    #df = pd.read_csv(ats_pt_f, delimiter = "\t",encoding='cp1252',error_bad_lines=False)
    df = pd.read_csv(ats_pt_f, delimiter = "\t",encoding='cp1252',on_bad_lines='skip')

    df.rename(columns={'Comments': 'ATS Comments'}, inplace=True)
    
    df.loc[(df['Accepted Date'].isnull()) & 
       (df['Rejected Date'].isnull()) & 
       (df['Submission Review Complete Date'].notnull()),
       'Accepted Date'] = df['Submission Review Complete Date']
    
    df['Decision-making Office Name'].fillna(df['Intake Office Name'], inplace=True)
    df.loc[df['Authorization Type'].str.contains('Aquaculture'), 
           'Decision-making Office Name'] = 'Aquaculture'
    
    df['Decision-making Office Name'] = df['Decision-making Office Name'].str.upper()
    
    
    
    ######## TEMPORARY: until using NANAIMO (DISTRICT) is confirmed by Auth team ###########
    df.loc[df['Decision-making Office Name'] == 'NANAIMO (DISTRICT)', 'Decision-making Office Name'] = 'NANAIMO'
    ######## TEMPORARY: until using NANAIMO (DISTRICT) is confirmed by Auth team ###########
    
    
    
    for index,row in df.iterrows():
        z_nbr = 7 - len(str(row['File Number']))
        df.loc[index, 'File Number'] = z_nbr * '0' + str(row['File Number'])
     
    # fill na Onhold time with 0
    df['Total On Hold Time'].fillna(0, inplace=True)
    

    #add on-hold cols
    df['Project Number'] = df['Project Number'].astype(str)
    df = pd.merge(df, df_onh, how='left', on='Project Number')
    
    #add bring-forward cols
    df = pd.merge(df, df_bfw, how='left', on='Project Number')
    
    for col in df:
        if 'Date' in col:
            df[col] =  pd.to_datetime(df[col],
                               infer_datetime_format=True,
                               errors = 'coerce').dt.date
        elif 'Unnamed' in col:
            df.drop(col, axis=1, inplace=True)
    
    else:
        pass
            

    return df


def import_titan (tnt_f):
    """Reads the Titan work ledger report into a df"""
    df = pd.read_excel(tnt_f,'TITAN_RPT009',
                       converters={'FILE NUMBER':str})
    
    tasks = ['NEW APPLICATION','REPLACEMENT APPLICATION','AMENDMENT','ASSIGNMENT']
    df = df.loc[df['TASK DESCRIPTION'].isin(tasks)]
    
    df.rename(columns={'COMMENTS': 'TANTALIS COMMENTS'}, inplace=True)
 
    del_col = ['ORG. UNIT','MANAGING AGENCY','BCGS','LEGAL DESCRIPTION',
              'FDISTRICT','ADDRESS LINE 1','ADDRESS LINE 2','ADDRESS LINE 3',
              'CITY','PROVINCE','POSTAL CODE','COUNTRY','STATE','ZIP CODE']
    
    for col in df:
        if 'DATE' in col:
            df[col] =  pd.to_datetime(df[col],
                                   infer_datetime_format=True,
                                   errors = 'coerce').dt.date
        elif 'Unnamed' in col:
            df.drop(col, axis=1, inplace=True)
        
        elif col in del_col:
            df.drop(col, axis=1, inplace=True)
            
        else:
            pass
            
    df.loc[df['PURPOSE'] == 'AQUACULTURE', 'DISTRICT OFFICE'] = 'AQUACULTURE'
    df.loc[df['DISTRICT OFFICE'] == 'COURTENAY', 'DISTRICT OFFICE'] = 'AQUACULTURE'
    df['DISTRICT OFFICE'] = df['DISTRICT OFFICE'].fillna(value='NANAIMO')
    
    return df


def calculate_metrics(df , grp_col, mtr_ids):
    """ Calculates Median and Mean metrics and return in df"""
    df_mtrs = []
    for mtr_id in mtr_ids:
        df_mtr = df.groupby(grp_col)[[mtr_id]].agg(['median', 'mean'])
        df_mtr.fillna(0, inplace=True)
        
        df_mtr.columns = [mtr_id+'_med',mtr_id+'_avg']
        
        df_mtr = df_mtr.reset_index()
        
        offices = ['AQUACULTURE','CAMPBELL RIVER','HAIDA GWAII',
                   'NANAIMO', 'PORT ALBERNI','PORT MCNEILL']
        
        if set(offices) != set(df_mtr['DISTRICT OFFICE'].unique()):
            new_rows = pd.DataFrame({'DISTRICT OFFICE': offices})
            df_mtr = pd.merge(new_rows, df_mtr, how='outer', on='DISTRICT OFFICE')
            df_mtr = df_mtr.fillna(0)
        else:
            df_mtr = df_mtr.sort_values(by='DISTRICT OFFICE')        
        
        
        df_mtr = pd.melt(df_mtr, id_vars=[grp_col])
        
        df_mtr = df_mtr.pivot_table(values='value', 
                                    index='variable', 
                                    columns=grp_col)
        
        vals = []
        for col in df_mtr.columns:
            vals.extend(df_mtr[col].to_list())
            
        mtr_cols = ['AQ avg','AQ med','CR avg','CR med',
                    'HG avg','HG med','NA avg','NA med',
                    'PA avg','PA med','PM avg','PM med']
        
        df_mtr = pd.DataFrame(data=[vals], columns=mtr_cols)

        df_mtr['WC avg'] = df.loc[df[mtr_id] != 0, mtr_id].mean()
        df_mtr['WC med'] = df.loc[df[mtr_id] != 0, mtr_id].median()

        df_mtr.fillna(0, inplace=True)        
        df_mtr = df_mtr.round().astype(int)
        
        df_mtr['METRIC ID'] = mtr_id
        
        df_mtrs.append(df_mtr)
    
    df_mtr = pd.concat(df_mtrs)
 
    
    return df_mtr


def create_rpt_01(rpt_date,df_tnt,df_ats):
    """ Creates Report 01- Files with FCBC"""
    ats_a = df_ats.loc[df_ats['Authorization Status'] == 'Active']
    #active = ats_a['File Number'].to_list()
    
    df_01= ats_a.loc[(ats_a['Received Date'].notnull()) &
                     (ats_a['Received Date'] <= rpt_date) &
                     (ats_a['Submission Review Complete Date'].isnull())]
    
    
    df_01 = pd.merge(df_01, df_tnt, how='left',
                     left_on='File Number',
                     right_on='FILE NUMBER')
    
    df_01= df_01.loc[(df_01['STATUS'].isnull())]
    
    df_01['Received Date'] = pd.to_datetime(df_01['Received Date']
                                             .fillna(pd.NaT), errors='coerce')
    
    df_01.sort_values(by=['Received Date'], ascending=False, inplace=True)
    df_01.reset_index(drop = True, inplace = True)
    
    df_01['DISTRICT OFFICE'] = df_01['Decision-making Office Name'] 
    
    df_01['Total On Hold Time'].fillna(0, inplace=True)
    
    #Calulcate metrics
    df_01_nw= df_01.loc[df_01['Authorization Type']!='Replacements']
    df_01_rp= df_01.loc[df_01['Authorization Type']=='Replacements']
    
    rpt_date = pd.to_datetime(rpt_date)

    for df in [df_01,df_01_nw,df_01_rp]:
        
        df['mtr01']  = (rpt_date - df['Received Date']).dt.days
        
    metrics = ['mtr01']
    df_01_mtr_nw = calculate_metrics(df_01_nw , 'DISTRICT OFFICE',metrics)
    df_01_mtr_rp = calculate_metrics(df_01_rp , 'DISTRICT OFFICE',metrics )
    

    return df_01,df_01_nw,df_01_rp,df_01_mtr_nw,df_01_mtr_rp


def create_rpt_02(rpt_date, df_tnt, df_ats):
    """Creates Report 02 - Files in Queue"""
    
    # Filter active records in df_ats
    ats_r = df_ats.loc[df_ats['Authorization Status'] == 'Active']
    active = ats_r['File Number'].to_list()
    
    # Filter records in df_tnt based on multiple conditions
    df_02 = df_tnt.loc[(df_tnt['TASK DESCRIPTION'].isin(['NEW APPLICATION', 'REPLACEMENT APPLICATION'])) &
                       (df_tnt['FILE NUMBER'].isin(active)) &
                       (df_tnt['OTHER EMPLOYEES ASSIGNED TO'].str.contains('WCR_', na=False) | 
                        df_tnt['OTHER EMPLOYEES ASSIGNED TO'].isnull()) &
                       (df_tnt['STATUS'] == 'ACCEPTED') &
                       (df_tnt['CREATED DATE'] <= rpt_date)]
    
    # Sort values
    df_02.sort_values(by='RECEIVED DATE', ascending=False, inplace=True)
    df_ats.sort_values(by='Received Date', ascending=False, inplace=True)
    
    # Calculate Join Start Date and Join End Date
    df_ats['Join Start Date'] = df_ats['Accepted Date'] - pd.DateOffset(months=6)
    df_ats['Join End Date'] = df_ats['Accepted Date'] + pd.DateOffset(months=6)
    
    # Add count columns for merging
    df_02['count'] = df_02.groupby('FILE NUMBER').cumcount()
    df_ats['count'] = df_ats.groupby('File Number').cumcount()
    
    # Merge df_02 with df_ats
    df_02 = pd.merge(df_02, df_ats, how='left',
                     left_on=['FILE NUMBER', 'count'],
                     right_on=['File Number', 'count'])
    
    # Check the CREATED DATE against Join Start Date and Join End Date
    for index, row in df_02.iterrows():
        # Ensure comparison between compatible types
        created_date = pd.to_datetime(row['CREATED DATE'])
        join_start_date = pd.to_datetime(row['Join Start Date'])
        join_end_date = pd.to_datetime(row['Join End Date'])
        
        if not (join_start_date <= created_date <= join_end_date):
            for col in df_ats.columns:
                df_02.at[index, col] = None
    
    # Reset index and sort values by CREATED DATE
    df_02.reset_index(drop=True, inplace=True)
    df_02.sort_values(by=['CREATED DATE'], ascending=False, inplace=True)
    
    # Handle missing values in Total On Hold Time
    df_02['Total On Hold Time'].fillna(0, inplace=True)
    
    # Calculate metrics
    rpt_date = pd.to_datetime(rpt_date)
    
    df_02['Submission Review Complete Date'] = pd.to_datetime(df_02['Submission Review Complete Date']
                                             .fillna(pd.NaT), errors='coerce')
    df_02['Received Date'] = pd.to_datetime(df_02['Received Date']
                                             .fillna(pd.NaT), errors='coerce')        
        
    # Separate DataFrames for NEW APPLICATION and REPLACEMENT APPLICATION
    df_02_nw = df_02.loc[df_02['TASK DESCRIPTION'] == 'NEW APPLICATION']
    df_02_rp = df_02.loc[df_02['TASK DESCRIPTION'] == 'REPLACEMENT APPLICATION']
    
    # Calculate metrics for each DataFrame
    for df in [df_02, df_02_nw, df_02_rp]:
        df['mtr02'] = (df['Submission Review Complete Date'] - df['Received Date']).dt.days
        df['mtr03'] = (rpt_date - df['Submission Review Complete Date']).dt.days
    
    metrics = ['mtr02', 'mtr03']
    df_02_mtr_nw = calculate_metrics(df_02_nw, 'DISTRICT OFFICE', metrics)
    df_02_mtr_rp = calculate_metrics(df_02_rp, 'DISTRICT OFFICE', metrics)
    
    return df_02, df_02_nw, df_02_rp, df_02_mtr_nw, df_02_mtr_rp


def create_rpt_03 (rpt_date,df_tnt,df_ats):
    """ Creates Report 03- Files in Active Review"""
    df_ats_h = df_ats.loc[df_ats['Authorization Status'] == 'On Hold']
    onhold= df_ats_h['File Number'].to_list()
    #df_ats = df_ats.loc[df_ats['Authorization Status'].isin(['Active','Closed'])]
   
    
    df_03= df_tnt.loc[((~df_tnt['OTHER EMPLOYEES ASSIGNED TO'].str.contains('WCR_',na=False)) & 
                       (df_tnt['OTHER EMPLOYEES ASSIGNED TO'].notnull())) &
                       (df_tnt['REPORTED DATE'].isnull()) &
                       (~df_tnt['FILE NUMBER'].isin(onhold)) &
                       (df_tnt['TASK DESCRIPTION'].isin(['NEW APPLICATION', 'REPLACEMENT APPLICATION'])) &            
                       (df_tnt['STATUS'] == 'ACCEPTED') &
                       (df_tnt['CREATED DATE'] <= rpt_date)]
 
    df_03.sort_values(by='RECEIVED DATE', ascending=False,inplace=True)
    df_ats.sort_values(by='Received Date', ascending=False,inplace=True)
    
    df_ats['Join Start Date'] = df_ats['Accepted Date'] - pd.DateOffset(months=6)
    df_ats['Join End Date'] = df_ats['Accepted Date'] + pd.DateOffset(months=6)
    
    df_03['count'] = df_03.groupby('FILE NUMBER').cumcount()
    df_ats['count'] = df_ats.groupby('File Number').cumcount()
    
    df_03 = pd.merge(df_03, df_ats, how='left',
                     left_on=['FILE NUMBER','count'],
                     right_on=['File Number','count'])
     
    for index, row in df_03.iterrows():
        # Ensure comparison between compatible types
        created_date = pd.to_datetime(row['CREATED DATE'])
        join_start_date = pd.to_datetime(row['Join Start Date'])
        join_end_date = pd.to_datetime(row['Join End Date'])
        
        if not (join_start_date <= created_date <= join_end_date):
            for col in df_ats.columns:
                df_03.at[index, col] = None
    
    df_03.sort_values(by=['CREATED DATE'], ascending=False, inplace=True)
    df_03.reset_index(drop = True, inplace = True)
    
    df_03['Total On Hold Time'].fillna(0, inplace=True)

    #Calulcate metrics
    rpt_date = pd.to_datetime(rpt_date)
    
    # for replacements only, use RECEIVED DATE instead of submisson review date to calculate mtr4
    df_03.loc[df_03['TASK DESCRIPTION'] == 'REPLACEMENT APPLICATION', 'Submission Review Complete Date'] = df_03['RECEIVED DATE']
    
    df_03['Bring Forward Date'] = pd.to_datetime(df_03['Bring Forward Date']
                                                 .fillna(pd.NaT), errors='coerce')
    df_03['Submission Review Complete Date'] = pd.to_datetime(df_03['Submission Review Complete Date']
                                                              .fillna(pd.NaT), errors='coerce')
    df_03['RECEIVED DATE'] = pd.to_datetime(df_03['RECEIVED DATE']
                                           .fillna(pd.NaT), errors='coerce')
    df_03['First Nation Start Date'] = pd.to_datetime(df_03['First Nation Start Date']
                                                      .fillna(pd.NaT), errors='coerce')
    df_03['First Nation Completion Date'] = pd.to_datetime(df_03['First Nation Completion Date']
                                                           .fillna(pd.NaT), errors='coerce')
    
    df_03_nw= df_03.loc[df_03['TASK DESCRIPTION']=='NEW APPLICATION']
    df_03_rp= df_03.loc[df_03['TASK DESCRIPTION']=='REPLACEMENT APPLICATION']    
    
    
    for df in [df_03,df_03_nw,df_03_rp]:
        df['mtr04'] = (df['Bring Forward Date'] - df['Submission Review Complete Date']).dt.days
        #df['mtr05'] = (rpt_date - df['First Nation Start Date']).dt.days
        df['mtr06'] = (df['First Nation Completion Date'] - df['First Nation Start Date']).dt.days
        df['mtr07'] = (rpt_date - df['Bring Forward Date']).dt.days

    metrics= ['mtr04','mtr06','mtr07']
    df_03_mtr_nw = calculate_metrics(df_03_nw , 'DISTRICT OFFICE', metrics ) 
    df_03_mtr_rp = calculate_metrics(df_03_rp , 'DISTRICT OFFICE', metrics ) 


    return df_03,df_03_nw,df_03_rp,df_03_mtr_nw,df_03_mtr_rp,onhold


def create_rpt_03_1 (rpt_date,df03):
    """ Creates Report 03-1- Files in Consultation"""
    df_031= df03.loc[(df03['First Nation Start Date'].notnull()) &
                     (df03['First Nation Completion Date'].isnull())]

    df_031.drop(['mtr04','mtr06','mtr07'], axis=1, inplace=True)
    
    df_031.sort_values(by=['First Nation Start Date'], ascending=False, inplace=True)
    df_031.reset_index(drop = True, inplace = True)
    
    df_031['Total On Hold Time'].fillna(0, inplace=True)

    #Calulcate metrics
    rpt_date = pd.to_datetime(rpt_date)
    
    # for replacements only, use RECEIVED DATE instead of submisson review date to calculate mtr4
    df_031.loc[df_031['TASK DESCRIPTION'] == 'REPLACEMENT APPLICATION', 'Submission Review Complete Date'] = df_031['RECEIVED DATE']
    

    df_031['First Nation Start Date'] = pd.to_datetime(df_031['First Nation Start Date'])

    
    df_031_nw= df_031.loc[df_031['TASK DESCRIPTION']=='NEW APPLICATION']
    df_031_rp= df_031.loc[df_031['TASK DESCRIPTION']=='REPLACEMENT APPLICATION']    
    
    
    for df in [df_031,df_031_nw,df_031_rp]:

        df['mtr05'] = (rpt_date - df['First Nation Start Date']).dt.days

    metrics= ['mtr05']
    df_031_mtr_nw = calculate_metrics(df_031_nw , 'DISTRICT OFFICE', metrics ) 
    df_031_mtr_rp = calculate_metrics(df_031_rp , 'DISTRICT OFFICE', metrics ) 


    return df_031,df_031_nw,df_031_rp,df_031_mtr_nw,df_031_mtr_rp


def create_rpt_04 (rpt_date,df_tnt,df_ats):
    """ Creates Report 04- Files Awaiting Decision"""
    df_ats = df_ats.loc[df_ats['Authorization Status'].isin(['Active','Closed'])]
    
    df_04= df_tnt.loc[(df_tnt['REPORTED DATE'].notnull()) &
                    (df_tnt['ADJUDICATED DATE'].isnull()) &
                    (df_tnt['TASK DESCRIPTION'].isin(['NEW APPLICATION', 'REPLACEMENT APPLICATION'])) &
                    (df_tnt['STATUS'] == 'ACCEPTED') &
                    (df_tnt['REPORTED DATE'] <= rpt_date)]

    df_04.sort_values(by='RECEIVED DATE', ascending=False,inplace=True)
    df_ats.sort_values(by='Received Date', ascending=False,inplace=True)
    
    df_ats['Join Start Date'] = df_ats['Accepted Date'] - pd.DateOffset(months=6)
    df_ats['Join End Date'] = df_ats['Accepted Date'] + pd.DateOffset(months=6)
    
    df_04['count'] = df_04.groupby('FILE NUMBER').cumcount()
    df_ats['count'] = df_ats.groupby('File Number').cumcount()
    
    df_04 = pd.merge(df_04, df_ats, how='left',
                     left_on=['FILE NUMBER','count'],
                     right_on=['File Number','count'])
    
    for index, row in df_04.iterrows():
        # Ensure comparison between compatible types
        created_date = pd.to_datetime(row['CREATED DATE'])
        join_start_date = pd.to_datetime(row['Join Start Date'])
        join_end_date = pd.to_datetime(row['Join End Date'])
        
        if not (join_start_date <= created_date <= join_end_date):
            for col in df_ats.columns:
                df_04.at[index, col] = None
    
    df_04.sort_values(by=['REPORTED DATE'], ascending=False, inplace=True)
    df_04.reset_index(drop = True, inplace = True)
    
    df_04['Total On Hold Time'].fillna(0, inplace=True)

    #Calulcate metrics
    rpt_date = pd.to_datetime(rpt_date)

    df_04['Bring Forward Date'] = pd.to_datetime(df_04['Bring Forward Date']
                                                 .fillna(pd.NaT), errors='coerce')

    df_04['REPORTED DATE'] = pd.to_datetime(df_04['REPORTED DATE']
                                                 .fillna(pd.NaT), errors='coerce') 
    
    df_04_nw= df_04.loc[df_04['TASK DESCRIPTION']=='NEW APPLICATION']
    df_04_rp= df_04.loc[df_04['TASK DESCRIPTION']=='REPLACEMENT APPLICATION']        
    
    
    for df in [df_04,df_04_nw,df_04_rp]:

        df['mtr08'] = (df['REPORTED DATE'] - df['Bring Forward Date']).dt.days
        df['mtr09'] = (rpt_date - df['REPORTED DATE']).dt.days

    metrics= ['mtr08','mtr09']
    df_04_mtr_nw = calculate_metrics(df_04_nw , 'DISTRICT OFFICE', metrics )  
    df_04_mtr_rp = calculate_metrics(df_04_rp , 'DISTRICT OFFICE', metrics ) 
    
    return df_04,df_04_nw,df_04_rp,df_04_mtr_nw,df_04_mtr_rp


def create_rpt_05 (rpt_date,df_tnt,df_ats):
    """ Creates Report 05- Files Awaiting Offer"""
    df_ats = df_ats.loc[df_ats['Authorization Status'].isin(['Active','Closed'])]
    
    df_05= df_tnt.loc[(df_tnt['ADJUDICATED DATE'].notnull()) &
                     (df_tnt['OFFERED DATE'].isnull()) &
                     (df_tnt['TASK DESCRIPTION'].isin(['NEW APPLICATION', 'REPLACEMENT APPLICATION'])) &            
                     (df_tnt['STATUS'] == 'ACCEPTED') &
                     (df_tnt['ADJUDICATED DATE'] <= rpt_date)]

    df_05.sort_values(by='RECEIVED DATE', ascending=False,inplace=True)
    df_ats.sort_values(by='Received Date', ascending=False,inplace=True)
    
    df_ats['Join Start Date'] = df_ats['Accepted Date'] - pd.DateOffset(months=6)
    df_ats['Join End Date'] = df_ats['Accepted Date'] + pd.DateOffset(months=6)
    
    df_05['count'] = df_05.groupby('FILE NUMBER').cumcount()
    df_ats['count'] = df_ats.groupby('File Number').cumcount()
    
    df_05 = pd.merge(df_05, df_ats, how='left',
                     left_on=['FILE NUMBER','count'],
                     right_on=['File Number','count'])
    
    for index, row in df_05.iterrows():
        # Ensure comparison between compatible types
        created_date = pd.to_datetime(row['CREATED DATE'])
        join_start_date = pd.to_datetime(row['Join Start Date'])
        join_end_date = pd.to_datetime(row['Join End Date'])
        
        if not (join_start_date <= created_date <= join_end_date):
            for col in df_ats.columns:
                df_05.at[index, col] = None
    
    df_05['ADJUDICATED DATE'] = pd.to_datetime(df_05['ADJUDICATED DATE']
                                                 .fillna(pd.NaT), errors='coerce')
    df_05['REPORTED DATE'] = pd.to_datetime(df_05['REPORTED DATE']
                                                 .fillna(pd.NaT), errors='coerce')
    
    df_05.sort_values(by=['ADJUDICATED DATE'], ascending=False, inplace=True)
    df_05.reset_index(drop = True, inplace = True)
    
    df_05['Total On Hold Time'].fillna(0, inplace=True)
    
    rpt_date = pd.to_datetime(rpt_date)

    #Calulcate metrics
    df_05_nw= df_05.loc[df_05['TASK DESCRIPTION']=='NEW APPLICATION']
    df_05_rp= df_05.loc[df_05['TASK DESCRIPTION']=='REPLACEMENT APPLICATION']  
    

    for df in [df_05,df_05_nw,df_05_rp]:
        df['mtr10'] = (df['ADJUDICATED DATE'] - df['REPORTED DATE']).dt.days
        df['mtr11'] = (rpt_date - df['ADJUDICATED DATE']).dt.days

    metrics= ['mtr10','mtr11']
    df_05_mtr_nw = calculate_metrics(df_05_nw , 'DISTRICT OFFICE', metrics )  
    df_05_mtr_rp = calculate_metrics(df_05_rp , 'DISTRICT OFFICE', metrics )  
    
    return df_05,df_05_nw,df_05_rp,df_05_mtr_nw,df_05_mtr_rp


def create_rpt_06 (rpt_date,df_tnt,df_ats):
    """ Creates Report 06- Files awaiting Offer Acceptance"""
    df_ats = df_ats.loc[df_ats['Authorization Status'].isin(['Active','Closed'])]
    df_06= df_tnt.loc[(df_tnt['OFFERED DATE'].notnull()) &
                      (df_tnt['OFFER ACCEPTED DATE'].isnull())&
                      (df_tnt['TASK DESCRIPTION'].isin(['NEW APPLICATION', 'REPLACEMENT APPLICATION'])) &                      
                      (df_tnt['STATUS'] == 'OFFERED') &
                      (df_tnt['OFFERED DATE'] <= rpt_date)]
    
    df_06.sort_values(by='RECEIVED DATE', ascending=False,inplace=True)
    df_ats.sort_values(by='Received Date', ascending=False,inplace=True)
    
    df_ats['Join Start Date'] = df_ats['Accepted Date'] - pd.DateOffset(months=6)
    df_ats['Join End Date'] = df_ats['Accepted Date'] + pd.DateOffset(months=6)
    
    df_06['count'] = df_06.groupby('FILE NUMBER').cumcount()
    df_ats['count'] = df_ats.groupby('File Number').cumcount()
    
    df_06 = pd.merge(df_06, df_ats, how='left',
                     left_on=['FILE NUMBER','count'],
                     right_on=['File Number','count'])
    
    for index, row in df_06.iterrows():
        # Ensure comparison between compatible types
        created_date = pd.to_datetime(row['CREATED DATE'])
        join_start_date = pd.to_datetime(row['Join Start Date'])
        join_end_date = pd.to_datetime(row['Join End Date'])
        
        if not (join_start_date <= created_date <= join_end_date):
            for col in df_ats.columns:
                df_06.at[index, col] = None
                
                
    df_06['OFFERED DATE'] = pd.to_datetime(df_06['OFFERED DATE']
                                            .fillna(pd.NaT), errors='coerce')
    df_06['ADJUDICATED DATE'] = pd.to_datetime(df_06['ADJUDICATED DATE']
                                            .fillna(pd.NaT), errors='coerce')
        
    df_06.sort_values(by=['OFFERED DATE'], ascending=False, inplace=True)
    df_06.reset_index(drop = True, inplace = True)
    
    df_06['Total On Hold Time'].fillna(0, inplace=True)

    rpt_date = pd.to_datetime(rpt_date)
    
    #Calulcate metrics
    df_06_nw= df_06.loc[df_06['TASK DESCRIPTION']=='NEW APPLICATION']
    df_06_rp= df_06.loc[df_06['TASK DESCRIPTION']=='REPLACEMENT APPLICATION']  

    for df in [df_06,df_06_nw,df_06_rp]:
        df['mtr12'] = (df['OFFERED DATE'] - df['ADJUDICATED DATE']).dt.days
        df['mtr13'] = (rpt_date - df['OFFERED DATE']).dt.days
    
    metrics= ['mtr12','mtr13']
    df_06_mtr_nw = calculate_metrics(df_06_nw , 'DISTRICT OFFICE', metrics )  
    df_06_mtr_rp = calculate_metrics(df_06_rp , 'DISTRICT OFFICE', metrics )  
    
    return df_06,df_06_nw,df_06_rp,df_06_mtr_nw,df_06_mtr_rp


def create_rpt_07 (rpt_date,df_tnt,df_ats):
    """ Creates Report 07- Files with Offer Accepted"""
    df_ats = df_ats.loc[df_ats['Authorization Status'].isin(['Active','Closed'])]
    df_07= df_tnt.loc[(df_tnt['OFFER ACCEPTED DATE'].notnull()) &
                     (df_tnt['TASK DESCRIPTION'].isin(['NEW APPLICATION', 'REPLACEMENT APPLICATION'])) &                      
                      (df_tnt['STATUS'] == 'OFFER ACCEPTED') &
                      (df_tnt['OFFER ACCEPTED DATE'] <= rpt_date)]
    
    df_ats['Join Start Date'] = df_ats['Accepted Date'] - pd.DateOffset(months=6)
    df_ats['Join End Date'] = df_ats['Accepted Date'] + pd.DateOffset(months=6)
    
    df_07['count'] = df_07.groupby('FILE NUMBER').cumcount()
    df_ats['count'] = df_ats.groupby('File Number').cumcount()
    
    df_07 = pd.merge(df_07, df_ats, how='left',
                     left_on=['FILE NUMBER','count'],
                     right_on=['File Number','count'])
    
    for index, row in df_07.iterrows():
        # Ensure comparison between compatible types
        created_date = pd.to_datetime(row['CREATED DATE'])
        join_start_date = pd.to_datetime(row['Join Start Date'])
        join_end_date = pd.to_datetime(row['Join End Date'])
        
        if not (join_start_date <= created_date <= join_end_date):
            for col in df_ats.columns:
                df_07.at[index, col] = None
    
    df_07['OFFERED DATE'] = pd.to_datetime(df_07['OFFERED DATE']
                                            .fillna(pd.NaT), errors='coerce')
    df_07['OFFER ACCEPTED DATE'] = pd.to_datetime(df_07['OFFER ACCEPTED DATE']
                                            .fillna(pd.NaT), errors='coerce')
    
    df_07.sort_values(by=['OFFER ACCEPTED DATE'], ascending=False, inplace=True)
    df_07.reset_index(drop = True, inplace = True)
    
    df_07['Total On Hold Time'].fillna(0, inplace=True)
    
    rpt_date = pd.to_datetime(rpt_date)

    #Calulcate metrics
    df_07_nw= df_07.loc[df_07['TASK DESCRIPTION']=='NEW APPLICATION']
    df_07_rp= df_07.loc[df_07['TASK DESCRIPTION']=='REPLACEMENT APPLICATION']  
    
    for df in [df_07,df_07_nw,df_07_rp]:
        df['mtr14'] = (df['OFFER ACCEPTED DATE'] - df['OFFERED DATE']).dt.days
        df['mtr15'] = (rpt_date - df['OFFER ACCEPTED DATE']).dt.days
    
    metrics= ['mtr14','mtr15']
    df_07_mtr_nw = calculate_metrics(df_07_nw , 'DISTRICT OFFICE', metrics )  
    df_07_mtr_rp = calculate_metrics(df_07_rp , 'DISTRICT OFFICE', metrics )  
    
    return df_07,df_07_nw,df_07_rp,df_07_mtr_nw,df_07_mtr_rp


def create_rpt_08 (rpt_date,df_tnt,df_ats):
    """ Creates Report 08- Files Completed"""
    df_ats = df_ats.loc[df_ats['Authorization Status'].isin(['Active','Closed'])]
    
    first_day_of_month = rpt_date.replace(day=1)
    
    df_08= df_tnt.loc[(df_tnt['COMPLETED DATE'].notnull()) &
                      (df_tnt['COMPLETED DATE'] >= first_day_of_month) &
                      (df_tnt['TASK DESCRIPTION'].isin(['NEW APPLICATION', 'REPLACEMENT APPLICATION'])) &
                      (df_tnt['STATUS'] == 'DISPOSITION IN GOOD STANDING') &
                      (df_tnt['COMPLETED DATE'] <= rpt_date)]
    
    df_ats['Join Start Date'] = df_ats['Accepted Date'] - pd.DateOffset(months=6)
    df_ats['Join End Date'] = df_ats['Accepted Date'] + pd.DateOffset(months=6)
    
    df_08['count'] = df_08.groupby('FILE NUMBER').cumcount()
    df_ats['count'] = df_ats.groupby('File Number').cumcount()
    
    df_08 = pd.merge(df_08, df_ats, how='left',
                     left_on=['FILE NUMBER','count'],
                     right_on=['File Number','count'])
    
    for index, row in df_08.iterrows():
        # Ensure comparison between compatible types
        created_date = pd.to_datetime(row['CREATED DATE'])
        join_start_date = pd.to_datetime(row['Join Start Date'])
        join_end_date = pd.to_datetime(row['Join End Date'])
        
        if not (join_start_date <= created_date <= join_end_date):
            for col in df_ats.columns:
                df_08.at[index, col] = None
    
    df_08['COMPLETED DATE'] = pd.to_datetime(df_08['COMPLETED DATE']
                                            .fillna(pd.NaT), errors='coerce')
    df_08['ADJUDICATED DATE'] = pd.to_datetime(df_08['ADJUDICATED DATE']
                                            .fillna(pd.NaT), errors='coerce')
    df_08['RECEIVED DATE'] = pd.to_datetime(df_08['RECEIVED DATE']
                                            .fillna(pd.NaT), errors='coerce')
    
    df_08.sort_values(by=['COMPLETED DATE'], ascending=False, inplace=True)
    df_08.reset_index(drop = True, inplace = True)
    
    df_08['Total On Hold Time'].fillna(0, inplace=True)

    #Calulcate metrics
    df_08_nw= df_08.loc[df_08['TASK DESCRIPTION']=='NEW APPLICATION']
    df_08_rp= df_08.loc[df_08['TASK DESCRIPTION']=='REPLACEMENT APPLICATION'] 
    

    for df in [df_08,df_08_nw,df_08_rp]:
        df['mtr16'] = (df['COMPLETED DATE'] - df['ADJUDICATED DATE']).dt.days
        df['mtr17'] = (df['COMPLETED DATE'] - df['RECEIVED DATE']).dt.days

    metrics= ['mtr16','mtr17']
    df_08_mtr_nw = calculate_metrics(df_08_nw , 'DISTRICT OFFICE', metrics ) 
    df_08_mtr_rp = calculate_metrics(df_08_rp , 'DISTRICT OFFICE', metrics )  

    
    return df_08,df_08_nw,df_08_rp,df_08_mtr_nw,df_08_mtr_rp


def create_rpt_09(rpt_date, df_tnt, df_ats):
    """Creates Report 09 - Files On Hold"""
    # Filter df_ats for relevant records
    df_ats = df_ats.loc[(df_ats['Authorization Status'] == 'On Hold') &
                        (df_ats['Accepted Date'].notnull()) &
                        (df_ats['On Hold Start Date'] <= rpt_date)]
    
    hold_l = df_ats['File Number'].to_list()
    
    # Filter df_tnt based on STATUS and FILE NUMBER
    df_09 = df_tnt.loc[(df_tnt['STATUS'] == 'ACCEPTED') & 
                       (df_tnt['FILE NUMBER'].isin(hold_l))]

    # Extract the year from the date columns
    df_09['CREATED DATE'] = pd.to_datetime(df_09['CREATED DATE'], errors='coerce')
    df_ats['Accepted Date'] = pd.to_datetime(df_ats['Accepted Date'], errors='coerce')
    
    df_09['tempo_join_date'] = df_09['CREATED DATE'].dt.year
    df_ats['tempo_join_date'] = df_ats['Accepted Date'].dt.year
    
    # Merge df_09 with df_ats
    df_09 = pd.merge(df_09, df_ats, how='left', left_on=['FILE NUMBER'], right_on=['File Number'])
    
    # Sort and reset index
    df_09.sort_values(by=['CREATED DATE'], ascending=False, inplace=True)
    df_09.reset_index(drop=True, inplace=True)
    
    df_09['Total On Hold Time'].fillna(0, inplace=True)

    # Calculate metrics
    rpt_date = pd.to_datetime(rpt_date)


    df_09['On Hold Start Date'] = pd.to_datetime(df_09['On Hold Start Date'].fillna(pd.NaT), errors='coerce')
    df_09['On Hold End Date'] = pd.to_datetime(df_09['On Hold End Date'].fillna(pd.NaT), errors='coerce') 

    df_09_nw = df_09.loc[df_09['TASK DESCRIPTION'] == 'NEW APPLICATION']
    df_09_rp = df_09.loc[df_09['TASK DESCRIPTION'] == 'REPLACEMENT APPLICATION'] 
    
    for df in [df_09, df_09_nw, df_09_rp]:
        df['mtr18'] = (rpt_date - df['On Hold Start Date']).dt.days

    metrics = ['mtr18']
    df_09_mtr_nw = calculate_metrics(df_09_nw, 'DISTRICT OFFICE', metrics)  
    df_09_mtr_rp = calculate_metrics(df_09_rp, 'DISTRICT OFFICE', metrics) 
    
    return df_09, df_09_nw, df_09_rp, df_09_mtr_nw, df_09_mtr_rp
    

def set_rpt_colums (dfs):
    """ Set the report columns"""
    cols = ['Region Name',
            'Business Area',
            'DISTRICT OFFICE',
            'FILE NUMBER',
            'Project Number',
            'STATUS',
            'TASK DESCRIPTION',
            'TYPE',
            'SUBTYPE',
            'PURPOSE',
            'SUBPURPOSE',
            'Authorization Type',
            'Authorization Status',
            'FCBC Assigned To',
            'OTHER EMPLOYEES ASSIGNED TO',
            'USERID ASSIGNED TO',
            'PRIORITY CODE',
            'CLIENT NAME',
            'LOCATION',
            'TANTALIS COMMENTS',
            'ATS Comments',
            'Received Date',
            'RECEIVED DATE',
            'Accepted Date',
            'CREATED DATE',
            'Submission Review Complete Date',
            'Bring Forward Date',
            'LAND STATUS DATE',
            'First Nation Start Date',
            'First Nation Completion Date',
            'FN Consultation Net Time',
            'REPORTED DATE',
            'ADJUDICATED DATE',
            'OFFERED DATE',
            'OFFER ACCEPTED DATE',
            'COMPLETED DATE',
            'On Hold Start Date',
            'On Hold End Date',
            'Reason For Hold',
            'Total On Hold Time',
            'Net Processing Time',
            'Total Processing Time']

    df_rpts = []   
    
    for df in dfs:

        for col in cols: 
            if 'Date' in col or 'DATE' in col:
                df[col] =  pd.to_datetime(df[col].fillna(pd.NaT),infer_datetime_format=True,
                                                                 errors = 'coerce').dt.date 
            if col not in df.columns:
                df[col] = pd.Series(dtype='object')
                

        df = df[cols+df.filter(regex='^mtr').columns.tolist()]
        df['Region Name'] = 'WEST COAST'
        df['Business Area'] = 'LANDS'

        df.rename({'Authorization Status': 'ATS STATUS', 
                   'STATUS': 'TANTALIS STATUS',
                   'TASK DESCRIPTION': 'APPLICATION TYPE',
                   'OTHER EMPLOYEES ASSIGNED TO':'FIELD EMPLOYEE',
                   'USERID ASSIGNED TO': 'EXAMINER NAME',
                   'Received Date': 'ATS RECEIVED DATE',
                   'RECEIVED DATE': 'TANTALIS RECEIVED DATE'}, 
                  axis=1, inplace=True)

        df.columns = [x.upper() for x in df.columns]
        
        df_rpts.append (df)
    
    return df_rpts
        

def create_summary_rpt (df_rpts):
    """Creates a summary  -Nbr of Files"""
    rpt_ids = ['rpt01','rpt02','rpt03','rpt03-1','rpt04',
               'rpt05','rpt06','rpt07','rpt08','rpt09']
    
    df_grs = []
    for df in df_rpts:
        df_gr = df.groupby('DISTRICT OFFICE')['REGION NAME'].count().reset_index()
        df_gr.sort_values(by=['DISTRICT OFFICE'], inplace = True)
        df_gr_pv = pd.pivot_table(df_gr, values='REGION NAME',
                        columns=['DISTRICT OFFICE'], fill_value=0)
        df_grs.append (df_gr_pv)
    
    df_sum_rpt = pd.concat(df_grs).reset_index(drop=True)
    df_sum_rpt.fillna(0, inplace=True)
    
    df_sum_rpt['WC files'] = df_sum_rpt.sum(axis=1)
    
    df_sum_rpt['REPORT ID'] = rpt_ids
    
    df_sum_rpt.rename({'AQUACULTURE': 'AQ files',
                       'CAMPBELL RIVER': 'CR files',
                       'HAIDA GWAII':'HG files',
                       'NANAIMO': 'NA files',
                       'PORT ALBERNI': 'PA files',
                       'PORT MCNEILL': 'PM files'}, axis=1, inplace=True)
    
    return df_sum_rpt,rpt_ids



def create_summary_mtr(df_mtrs):
    """Creates a summary- Nbr of Days"""


    df_sum_mtr = pd.concat(df_mtrs)
    df_sum_mtr = df_sum_mtr.reset_index(drop=True)
    

    return df_sum_mtr


def create_summary_all(template,df_sum_rpt,df_sum_mtr):
    """Create a Summary of Nbr files and days"""
    df_tmp = pd.read_excel(template)
    
    df_sum_all = pd.merge(df_tmp,df_sum_rpt,
                          how='left',
                          on='REPORT ID')
    
    df_sum_all = pd.merge(df_sum_all,df_sum_mtr,
                          how='left',
                          on='METRIC ID')

    
    sum_cols = ['REPORT ID', 'REPORT NAME', 'METRIC ID', 'METRIC NAME', 'WC files',
                'WC avg', 'WC med', 'AQ files','AQ avg', 'AQ med','CR files','CR avg', 
                'CR med','HG files','HG avg','HG med','NA files','NA avg', 'NA med',
                'PA files','PA avg', 'PA med','PM files','PM avg', 'PM med']
        
    df_sum_all= df_sum_all[sum_cols]  
    
    return df_sum_all


def analysis_tables (tmplt_anlz,df_sum_rpt,df_sum_mtr):
    """Create Analysis tables"""
    df_tmp= pd.read_excel(tmplt_anlz)

    df_anz_tim= pd.merge(df_tmp,df_sum_mtr[['METRIC ID','WC avg','WC med']],
                         how= 'left', on='METRIC ID')
    
    tm_stg= ['mtr01','mtr03','mtr07','mtr05','mtr09','mtr11','mtr13','mtr15','mtr18']
    tm_prc= ['mtr02','mtr04','mtr08','mtr06','mtr10','mtr12','mtr14','mtr16']
    
    df_anz_tim.loc[df_anz_tim['METRIC ID'].isin(tm_stg),'METRIC ID']='Time of Files at Stage'
    df_anz_tim.loc[df_anz_tim['METRIC ID'].isin(tm_prc),'METRIC ID']='Processing Time'
    
    df_anz_tim.rename(columns={'WC avg': 'Average','WC med': 'Median'}, inplace=True)

    df_anz_tim = df_anz_tim.pivot(index='REPORT NAME', columns='METRIC ID')
    df_anz_tim.columns = [f'{col[0]}_{col[1]}' for col in df_anz_tim.columns]
    df_anz_tim=df_anz_tim.reset_index()
    
    df_anz_tim.drop('REPORT ID_Processing Time', axis=1, inplace=True)
    df_anz_tim.rename(columns={'REPORT ID_Time of Files at Stage': 'REPORT ID'}, inplace=True)
    
    df_anz_tim= pd.merge(df_anz_tim,df_sum_rpt[['REPORT ID','WC files']],
                    how= 'left', on='REPORT ID')
    df_anz_tim.rename(columns={'WC files': '# Files at Stage',
                               'REPORT NAME': 'Stage'}, inplace=True)  
    
    df_anz_tim.sort_values('REPORT ID', ascending=True, inplace=True)
    
    df_anz_tim = df_anz_tim[['Stage','# Files at Stage','Average_Time of Files at Stage',
                     'Median_Time of Files at Stage','Average_Processing Time',
                     'Median_Processing Time']]
    
    df_anz_tim=df_anz_tim.reset_index(drop= True)
    
    
    df_tmp.drop_duplicates('REPORT ID', inplace= True)
    df_tmp.drop('METRIC ID', axis=1, inplace=True)
    df_tmp.loc[len(df_tmp)] = ['rpt08', 'Files Completed']
    df_tmp.sort_values('REPORT ID', ascending=True, inplace=True)
    
    cols= ['REPORT ID','AQ files', 'CR files', 'HG files', 'NA files', 'PA files', 'PM files']
    df_anz_off= pd.merge(df_tmp,df_sum_rpt[cols],
                         how= 'left', on='REPORT ID')
    
    df_anz_off.rename(columns={'REPORT NAME': 'Stage',
                               'AQ files': 'Aquaculture', 
                               'NA files': 'Nanaimo',
                               'HG files': 'HGNRD',
                               'CR files': 'CRNRD',
                               'PA files': 'SINRD',
                               'PM files': 'NICCNRD'}, inplace=True)
    
    df_anz_off.drop('REPORT ID', axis=1, inplace=True)
    df_anz_off = df_anz_off[['Stage','Nanaimo','Aquaculture','SINRD',
                             'CRNRD','NICCNRD','HGNRD']]
    
    return df_anz_tim,df_anz_off


def add_analysis_tables(df_anz_tim, df_anz_off, nw_rp, out_folder, filename):
    """Adds the Executive Summaries to the Main report"""
    out_file = os.path.join(out_folder, f"{filename}.xlsx")
    
    # Load the existing workbook
    workbook = openpyxl.load_workbook(out_file)
    
    if nw_rp == 'NEW':
        sheet_index = 0
        tab_tim_nme = 'Table3000'
        tab_off_nme = 'Table3001'
    else:
        sheet_index = 1
        tab_tim_nme = 'Table3002'
        tab_off_nme = 'Table3003'
    
    sheet_name = workbook.sheetnames[sheet_index]
    worksheet = workbook[sheet_name]
    
    # Write df_anz_tim to the worksheet
    start_row = 21
    start_col = 2  # Column B
    for r_idx, row in enumerate(dataframe_to_rows(df_anz_tim, index=False, header=True)):
        for c_idx, value in enumerate(row):
            worksheet.cell(row=start_row+r_idx, column=start_col+c_idx, value=value)
    
    # Write df_anz_off to the worksheet
    start_row_off = start_row + len(df_anz_tim) + 3
    for r_idx, row in enumerate(dataframe_to_rows(df_anz_off, index=False, header=True)):
        for c_idx, value in enumerate(row):
            worksheet.cell(row=start_row_off+r_idx, column=start_col+c_idx, value=value)
    
    # Add tables
    end_row_tim = start_row + len(df_anz_tim)
    end_col_tim = start_col + len(df_anz_tim.columns) - 1
    tab_tim = Table(displayName=tab_tim_nme, ref=f"{openpyxl.utils.get_column_letter(start_col)}{start_row}:{openpyxl.utils.get_column_letter(end_col_tim)}{end_row_tim}")
    
    end_row_off = start_row_off + len(df_anz_off)
    end_col_off = start_col + len(df_anz_off.columns) - 1
    tab_off = Table(displayName=tab_off_nme, ref=f"{openpyxl.utils.get_column_letter(start_col)}{start_row_off}:{openpyxl.utils.get_column_letter(end_col_off)}{end_row_off}")
    
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=True,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    tab_tim.tableStyleInfo = style
    tab_off.tableStyleInfo = style
    
    worksheet.add_table(tab_tim)
    worksheet.add_table(tab_off)
    
    # Save the workbook
    workbook.save(out_file)
    
    
def compute_chart (df, title_tag, out_folder, figname):
    """Computes a barplot of number of # of files and processing times """

    fig = px.bar(df, x='Stage', y='# Files at Stage',template='plotly')
    fig.update_traces(texttemplate='<b>%{y}</b>', textposition='auto')
    
    fig.add_trace(go.Scatter(x=df['Stage'], y=df['Average_Time of Files at Stage'], 
                             mode='markers', name='Average Time', 
                             marker=dict(symbol='x', color='red', size=12), yaxis='y2'))
    fig.add_trace(go.Scatter(x=df['Stage'], y=df['Median_Time of Files at Stage'], 
                             mode='markers', name='Median Time', 
                             marker=dict(symbol='circle', color='orange', size=12), yaxis='y2'))
    

    exld= ['Files in Consultation (with LO)','Files On Hold']
    df_sum= df.loc[~df['Stage'].isin(exld)]
    nbr_files= int(df_sum['# Files at Stage'].sum())
    title= """WCR Lands Applications Workflow Status - {} <br>[{} Total Files in Process, excl On Hold]
           """.format(title_tag, nbr_files)
    
    fig.update_layout(
        title=title,
        title_x=0.5,  
        yaxis=dict(title='# Files at Stage'),
        yaxis2=dict(title='Time (Days)', overlaying='y', side='right'),
        legend=dict(orientation='h', yanchor='top', y=1.06, xanchor='center', x=0.87)
    )
    
    out_chart= os.path.join('{}'.format(out_folder), figname+'.png')
    fig.write_image(out_chart, width=1200, height=800, scale=2)



def add_charts(nw_rp, out_folder, filename, figname):
    """ "Adds the charts to the Main report """
    
    out_file = os.path.join(out_folder, f"{filename}.xlsx")
    workbook = load_workbook(out_file)
    
    if nw_rp == 'NEW':
        sheet_index = 0
    else:
        sheet_index = 1

    sheet_name = workbook.sheetnames[sheet_index]
    worksheet = workbook[sheet_name]
    
    image_path = os.path.join(out_folder, f"{figname}.png")

    pil_image = PILImage.open(image_path)
    
    width_cm = 27
    height_cm = 17
    
    dpi = 96 
    width_px = int(width_cm * dpi / 2.54)
    height_px = int(height_cm * dpi / 2.54)
    
    resized_image = pil_image.resize((width_px, height_px))
    
    resized_image.save(image_path, "PNG")
    img = Image(image_path)

    cell = "J22"
    
    worksheet.add_image(img, cell)
    
    workbook.save(out_file)


def create_hitlists (df_rpts):
    mtr_lst= ['mtr01','mtr03','mtr07','mtr05','mtr09',
              'mtr11','mtr13','mtr15','mtr17','mtr18']
    
    dfs_htlst = []
    
    for i, df in enumerate(df_rpts):
        mtr= mtr_lst[i]
        df.sort_values(by=mtr.upper(), ascending=False, inplace=True)
        dfs_htlst.append(df.head(10))
        
    return dfs_htlst
    
  
def create_report (df_list, sheet_list,out_folder,filename):
    """ Exports dataframes to multi-tab excel spreasheet"""
    out_file= os.path.join('{}'.format(out_folder), filename+'.xlsx')
    writer = pd.ExcelWriter(out_file,engine='xlsxwriter')

    for dataframe, sheet in zip(df_list, sheet_list):
        dataframe = dataframe.reset_index(drop=True)
        dataframe.index = dataframe.index + 1

        dataframe.to_excel(writer, sheet_name=sheet, index=False, startrow=0 , startcol=0)

        worksheet = writer.sheets[sheet]
        #workbook = writer.book
        
        if sheet in ['Summary - NEW Applics','Summary - REP Applics']:
            worksheet.set_column(0, 0, 11)
            worksheet.set_column(1, 1, 27)
            worksheet.set_column(2, 2, 11)
            worksheet.set_column(3, 3, 37)
            worksheet.set_column(4, dataframe.shape[1], 10)
        
        else:
            worksheet.set_column(0, dataframe.shape[1], 20)

        col_names = [{'header': col_name} for col_name in dataframe.columns[:]]

        worksheet.add_table(0, 0, dataframe.shape[0], dataframe.shape[1]-1, 
                            {'columns': col_names})

    #writer.save()
    writer.close()    



def add_readme_page(readme_xlsx, out_folder, filename):
    source_workbook = openpyxl.load_workbook(readme_xlsx)
    source_sheet = source_workbook['README']
    
    rpt_xlsx = os.path.join('{}'.format(out_folder), filename+'.xlsx')
    target_workbook = openpyxl.load_workbook(rpt_xlsx)
    
    target_sheet = target_workbook.create_sheet(title=source_sheet.title, index=0)
    
    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = target_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            if cell.has_style:
                target_cell.font = cell.font.copy()
                target_cell.border = cell.border.copy()
                target_cell.fill = cell.fill.copy()
                target_cell.number_format = cell.number_format
                target_cell.protection = cell.protection.copy()
                target_cell.alignment = cell.alignment.copy()

    for i, col in enumerate(source_sheet.columns):
        source_width = source_sheet.column_dimensions[col[0].column_letter].width
        target_sheet.column_dimensions[get_column_letter(i+1)].width = source_width
    
    target_workbook.active = 0
    
    target_workbook.save(rpt_xlsx)
                



if __name__ == "__main__":
    
    print ('\nConnecting to BCGW.')
    '''
    hostname = 'bcgw.bcgov/idwprod1.bcgov'
    bcgw_user = os.getenv('bcgw_user')
    bcgw_pwd = os.getenv('bcgw_pwd')
    connection = connect_to_DB (bcgw_user,bcgw_pwd,hostname)
    '''
    wks= r'W:\lwbc\visr\Workarea\moez_labiadh\FILE_TRACKING'
    
    # The first day of previous month. Will be used to calculate Metrics
    today = date.today()
    first_day_month = today.replace(day=1)
    rpt_date= first_day_month - timedelta(days=1)
    
    rpt_month_str = rpt_date.strftime("%b%Y").lower()
    
    
    print ('\nImporting Input files')
    
    ats_date_tag= first_day_month.strftime("%Y%m%d")
    print('...TITAN workledger spreadsheet')
    tnt_f = os.path.join(wks,'00_INPUTS/TITAN_RPT009.xlsx')
    df_tnt = import_titan (tnt_f)
    
    print ('...ATS on-hold spreadsheet')
    ats_oh_f = os.path.join(wks, f'00_INPUTS/{ats_date_tag}_ats_oh.xls')
    df_onh= import_ats_oh (ats_oh_f)
    
    print ('...ATS bring-forward spreadsheet')
    ats_bf_f = os.path.join(wks, f'00_INPUTS/{ats_date_tag}_ats_bf.xls')
    df_bfw= import_ats_bf (ats_bf_f)
    
    print('...ATS processing time spreadsheet')
    ats_pt_f = os.path.join(wks, f'00_INPUTS/{ats_date_tag}_ats_pt.xls')
    df_ats = import_ats_pt (ats_pt_f, df_onh,df_bfw)
    
    print('\nComputing Reports.')
    dfs = []
    dfs_nw = []
    dfs_rp = []
    df_mtrs_nw = []
    df_mtrs_rp = []
        
    print('...report 01')
    df_01,df_01_nw,df_01_rp,df_01_mtr_nw,df_01_mtr_rp= create_rpt_01 (rpt_date,df_tnt,df_ats)
    dfs.append(df_01)
    dfs_nw.append(df_01_nw)
    dfs_rp.append(df_01_rp)
    df_mtrs_nw.append(df_01_mtr_nw)
    df_mtrs_rp.append(df_01_mtr_rp)
    
    print('...report 02')
    df_02,df_02_nw,df_02_rp,df_02_mtr_nw,df_02_mtr_rp = create_rpt_02 (rpt_date,df_tnt,df_ats)
    dfs.append(df_02)
    dfs_nw.append(df_02_nw)
    dfs_rp.append(df_02_rp)
    df_mtrs_nw.append(df_02_mtr_nw)
    df_mtrs_rp.append(df_02_mtr_rp)
    
    print('...report 03')
    df_03,df_03_nw,df_03_rp,df_03_mtr_nw,df_03_mtr_rp,onhold = create_rpt_03 (rpt_date,df_tnt,df_ats)
    dfs.append(df_03)
    dfs_nw.append(df_03_nw)
    dfs_rp.append(df_03_rp)
    df_mtrs_nw.append(df_03_mtr_nw)
    df_mtrs_rp.append(df_03_mtr_rp)
    
    print('...report 03-1')
    df_031,df_031_nw,df_031_rp,df_031_mtr_nw,df_031_mtr_rp= create_rpt_03_1 (rpt_date,df_03)
    dfs.append(df_031)
    dfs_nw.append(df_031_nw)
    dfs_rp.append(df_031_rp)
    df_mtrs_nw.append(df_031_mtr_nw)
    df_mtrs_rp.append(df_031_mtr_rp)
    
    print('...report 04')
    df_04,df_04_nw,df_04_rp,df_04_mtr_nw,df_04_mtr_rp = create_rpt_04(rpt_date,df_tnt,df_ats)
    dfs.append(df_04)
    dfs_nw.append(df_04_nw)
    dfs_rp.append(df_04_rp)
    df_mtrs_nw.append(df_04_mtr_nw)
    df_mtrs_rp.append(df_04_mtr_rp)
    
    print('...report 05')
    df_05,df_05_nw,df_05_rp,df_05_mtr_nw,df_05_mtr_rp = create_rpt_05 (rpt_date,df_tnt,df_ats)
    dfs.append(df_05)
    dfs_nw.append(df_05_nw)
    dfs_rp.append(df_05_rp)
    df_mtrs_nw.append(df_05_mtr_nw)
    df_mtrs_rp.append(df_05_mtr_rp)
    
    print('...report 06')
    df_06,df_06_nw,df_06_rp,df_06_mtr_nw,df_06_mtr_rp = create_rpt_06 (rpt_date,df_tnt,df_ats)
    dfs.append(df_06)
    dfs_nw.append(df_06_nw)
    dfs_rp.append(df_06_rp)
    df_mtrs_nw.append(df_06_mtr_nw)
    df_mtrs_rp.append(df_06_mtr_rp)
    
    print('...report 07')
    df_07,df_07_nw,df_07_rp,df_07_mtr_nw,df_07_mtr_rp = create_rpt_07 (rpt_date,df_tnt,df_ats)
    dfs.append(df_07)
    dfs_nw.append(df_07_nw)
    dfs_rp.append(df_07_rp)
    df_mtrs_nw.append(df_07_mtr_nw)
    df_mtrs_rp.append(df_07_mtr_rp)
    
    print('...report 08')
    df_08,df_08_nw,df_08_rp,df_08_mtr_nw,df_08_mtr_rp= create_rpt_08 (rpt_date,df_tnt,df_ats)
    dfs.append(df_08)
    dfs_nw.append(df_08_nw)
    dfs_rp.append(df_08_rp)
    df_mtrs_nw.append(df_08_mtr_nw)
    df_mtrs_rp.append(df_08_mtr_rp)
    
    print('...report 09')
    df_09,df_09_nw,df_09_rp,df_09_mtr_nw,df_09_mtr_rp = create_rpt_09 (rpt_date,df_tnt,df_ats)
    dfs.append(df_09)
    dfs_nw.append(df_09_nw)
    dfs_rp.append(df_09_rp)
    df_mtrs_nw.append(df_09_mtr_nw)
    df_mtrs_rp.append(df_09_mtr_rp)
    
    print('\nFormatting Reports')
    df_rpts = set_rpt_colums (dfs)
    df_rpts_nw = set_rpt_colums (dfs_nw)
    df_rpts_rp = set_rpt_colums (dfs_rp)
    
    print('\nCalculating Summary Stats')
    df_sum_rpt_nw,rpt_ids = create_summary_rpt (df_rpts_nw)
    df_sum_rpt_rp,rpt_ids = create_summary_rpt (df_rpts_rp)
    
    df_sum_mtr_nw= create_summary_mtr(df_mtrs_nw)
    df_sum_mtr_rp= create_summary_mtr(df_mtrs_rp)
    
    print ('\nCreating Analysis tables')
    tmplt_anlz = os.path.join(wks,'00_TEMPLATE/anz_template.xlsx')
    df_anz_tim_nw, df_anz_off_nw= analysis_tables (tmplt_anlz,df_sum_rpt_nw,df_sum_mtr_nw)
    df_anz_tim_rp, df_anz_off_rp= analysis_tables (tmplt_anlz,df_sum_rpt_rp,df_sum_mtr_rp)
    
    
    template = os.path.join(wks,'00_TEMPLATE/rpt_template.xlsx')
    df_sum_all_nw= create_summary_all(template,df_sum_rpt_nw,df_sum_mtr_nw)
    df_sum_all_rp= create_summary_all(template,df_sum_rpt_rp,df_sum_mtr_rp)
    
    # set the first 3 rows of the replacement summary to N/A.
    rows_range = slice(0, 3)
    cols_range = slice(4, 26)
    df_sum_all_rp.iloc[rows_range, cols_range] = 'n/a'
    
    print('\nCreating an Output folder')
    out_folder = os.path.join(wks, rpt_month_str)
    
    if not os.path.exists(out_folder):
        os.makedirs(out_folder)
        
    print('\nExporting the Main Report')
    df_list = [df_sum_all_nw,df_sum_all_rp] + df_rpts 
    sheet_list = ['Summary - NEW Applics','Summary - REP Applics'] + rpt_ids
    
    outfile_main_rpt = rpt_month_str + '_landFiles_tracker'
    
    create_report (df_list, sheet_list,out_folder,outfile_main_rpt)
    
    nw_rp= 'NEW'
    add_analysis_tables(df_anz_tim_nw,df_anz_off_nw,nw_rp,out_folder,outfile_main_rpt)
    nw_rp= 'REP'
    add_analysis_tables(df_anz_tim_rp,df_anz_off_rp,nw_rp,out_folder,outfile_main_rpt)
    
    
    print ('\nExporting the Hitlists Report')
    dfs_htlst= create_hitlists (df_rpts)
    
    dfs_htlst_lbls= ['hitlist_rpt01','hitlist_rpt02','hitlist_rpt03','hitlist_rpt03-1',
                     'hitlist_rpt04','hitlist_rpt05','hitlist_rpt06','hitlist_rpt07',
                     'hitlist_rpt08','hitlist_rpt09']
    
    outfile_hit = rpt_month_str + '_landFiles_tracker_hitlists'
    create_report (dfs_htlst, dfs_htlst_lbls, out_folder, outfile_hit)
    
    
    print ('\nComputing Charts')
    figname_nw= rpt_month_str+'_chart_processingTimes_new'
    title_tag_nw= 'New Files'
    compute_chart (df_anz_tim_nw, title_tag_nw, out_folder, figname_nw)
    
    figname_rp= rpt_month_str+'_chart_processingTimes_rep'
    title_tag_rp= 'Replacement Files'
    compute_chart (df_anz_tim_rp, title_tag_rp, out_folder, figname_rp)
    
    nw_rp= 'NEW'
    add_charts(nw_rp, out_folder, outfile_main_rpt,figname_nw)
    nw_rp= 'REP'
    add_charts(nw_rp, out_folder, outfile_main_rpt,figname_rp)
    
    readme_xlsx= os.path.join(wks,'00_TEMPLATE/readme_template.xlsx')
    add_readme_page(readme_xlsx, out_folder, outfile_main_rpt)
    