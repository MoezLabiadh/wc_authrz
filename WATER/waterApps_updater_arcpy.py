import os
import arcpy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


def update_wtshd_info(f, wtsh_fc,df_lk,saveOp):
    """Updates the Watershed and Region info of the Water Apps ledger"""
    wb = load_workbook(f)
    ws = wb['Active Applications']

    for i, row in enumerate(ws.iter_rows(min_row=2,values_only=True)):
        row_id = i+2

        if ws['X{}'.format(row_id)].value is None:
            print ("\n********Working on Row {}********".format (row_id))
            lat = row[10]
            long = row[11]

            print ('Latitude: {}'. format(lat))
            print ('Longitude: {}'. format(long))

            if (lat is None) or (long is None):
                raise Exception("Latitude or/and Longitude values are not specified!")

            elif not min(54.465, 48.2) < lat < max(54.5, 48.134):
                raise Exception("Latitude value is out of range!")


            elif not min(-122.634, -133.257) < long < max(-122.7, -133.3):
                    raise Exception("Longitude value is out of range!")

            else:
                pt = arcpy.Point(float(long),float(lat))
                sr = arcpy.SpatialReference("WGS 1984")
                ptGeo = arcpy.PointGeometry(pt,sr)

                pt_fc = 'in_memory\pt_fc'
                arcpy.management.CopyFeatures(ptGeo, pt_fc)

                intr = 'in_memory\_intersect'
                arcpy.Intersect_analysis([pt_fc,wtsh_fc], intr)

                with arcpy.da.UpdateCursor(intr,['WATER_LICENSING_WATERSHED_NAME']) as cursor:
                    for row in cursor:
                        wtrsh_name = row[0]

                ws['X{}'.format(row_id)] = wtrsh_name
                ws['Y{}'.format(row_id)]  = """=VLOOKUP(X{},'Pick Lists'!$L$1:$M$107,2,FALSE)""".format(row_id)

                df_lk_i = df_lk.loc[df_lk['WATER_LICENSING_WATERSHED'] == wtrsh_name]
                print ('Watershed: {}'.format(wtrsh_name))
                print ('Sub-region: {}'.format(df_lk_i['REGION'].iloc[0]))

                arcpy.Delete_management('in_memory')

        else:
            #raise Exception ('No Empty watershed values found in spreadsheet')
            pass

    #Add data validation for Watershed name.
    dv = DataValidation(type="list",
                        formula1= "'Pick Lists'!$L$2:$L$107",
                        allow_blank=False,
                        showDropDown= False)

    dv.add("X2:X" + str(ws.max_row))
    ws.add_data_validation(dv)

    if saveOp == 'Yes':
        print('Writing Watershed and Subregion info to the Spreadsheet')
        wb.save(f)

    else:
        pass

def main():
    workspace = r'\\spatialfiles.bcgov\Work\lwbc\visr\Workarea\moez_labiadh\WORKSPACE\20220916_waterLicencing_support'
    f = os.path.join (workspace,'Water Application Ledger_workingCopy2.xlsx')
    wtsh_fc = os.path.join (workspace,'data.gdb', 'watersheds_WC_all')

    saveOp = 'No' # yes or no


    df_lk = pd.read_excel(f, 'Pick Lists')
    df_lk = df_lk[['WATER_LICENSING_WATERSHED', 'REGION']]

    print ("Retrieving Watershed and Sub-region info...")
    update_wtshd_info(f, wtsh_fc,df_lk,saveOp)

main()
